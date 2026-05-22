import io
import re
from typing import Callable
from openpyxl import load_workbook

NAME_HEADER_PATTERN = re.compile(r'^(?:[おご]?(?:名前|氏名)|姓名|name)$', re.IGNORECASE)


def _find_name_columns(sheet):
    name_cols = set()
    first_row = next(sheet.iter_rows(values_only=True), None)
    if first_row is None:
        return name_cols
    for i, cell in enumerate(first_row):
        if isinstance(cell, str) and NAME_HEADER_PATTERN.search(cell.strip()):
            name_cols.add(i)
    return name_cols


def process_excel(content: bytes, redact: Callable[[str], str]) -> bytes:
    workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
    for sheet in workbook.worksheets:
        name_cols = _find_name_columns(sheet)
        for row_index, row in enumerate(sheet.iter_rows(values_only=False)):
            for col_index, cell in enumerate(row):
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                if row_index > 0 and col_index in name_cols and cell.value.strip():
                    cell.value = "[NAME]"
                else:
                    cell.value = redact(cell.value)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()
